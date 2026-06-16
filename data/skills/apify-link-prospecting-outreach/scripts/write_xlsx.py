#!/usr/bin/env python3
"""Write the final 30-column xlsx + run_metadata_final.json sidecar.

Usage:
  python3 scripts/write_xlsx.py --config campaign.json

Reads:
  {base}_drafted.json
  {base}_run_metadata.json   (optional — runner emits this; preserves runId/datasetIds)
Writes:
  {base}.xlsx                 (2 sheets: Outreach + Skipped)
  {base}_run_metadata_final.json
"""
import argparse
import json
import re
import sys
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

COLUMNS = [
    ("SERP Position", 12), ("Source Engines", 32), ("Keyword", 30),
    ("Article Title", 60), ("Article URL", 70), ("Domain", 24),
    ("Domain DR", 10), ("Page Traffic", 14), ("Referring Domains", 14),
    ("Prospect Tier", 12), ("Why This Prospect", 40), ("Article Author", 24),
    ("Author Source", 14), ("Publish Date", 14), ("Contact Full Name", 24),
    ("Contact Job Title", 30), ("Department", 16), ("Seniority", 14),
    ("Contact Email", 30), ("Email Verification", 16), ("Contact LinkedIn", 50),
    ("Company", 24), ("Outreach Type", 24), ("Partnership Offer", 30),
    ("Placement Source Sentence", 80), ("Placement With Link", 80),
    ("Placement New Insertion", 80), ("Suggested Email Copy", 100),
    ("Outreach Status", 16), ("Notes", 60),
]

SKIP_COLUMNS = [
    ("Domain", 28), ("Article URL", 70), ("Article Title", 60),
    ("Skip Reason", 60), ("Source Engines", 32), ("Why This Prospect", 40),
]


def extract_skip_reason(row):
    m = re.search(r'SKIP[^:]*:\s*([^|]+)', row.get("Notes", "") or "")
    return m.group(1).strip() if m else "(unknown)"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    args = ap.parse_args()

    cfg = json.load(open(args.config))
    base = cfg["base"]
    rows = json.load(open(f"{base}_drafted.json"))

    active = [r for r in rows if r["Outreach Status"] != "Skip"]
    skipped = [r for r in rows if r["Outreach Status"] == "Skip"]

    def active_sort_key(r):
        t = {"A": 0, "B": 1, "C": 2, "-": 3}.get(r["Prospect Tier"], 4)
        dr = -(r["Domain DR"] if isinstance(r["Domain DR"], (int, float)) else 0)
        sp = r["SERP Position"] if isinstance(r["SERP Position"], int) else 999
        return (t, dr, sp)

    active = sorted(active, key=active_sort_key)
    skipped = sorted(skipped, key=lambda r: (extract_skip_reason(r), r.get("Domain", "")))

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    tier_fills = {
        "A": PatternFill("solid", fgColor="D1FAE5"),
        "B": PatternFill("solid", fgColor="FEF3C7"),
        "C": PatternFill("solid", fgColor="F3F4F6"),
        "-": PatternFill("solid", fgColor="F3F4F6"),
    }
    skip_fill = PatternFill("solid", fgColor="FEE2E2")

    ws = wb.active
    ws.title = "Outreach"
    ws.freeze_panes = "A2"
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    for row_idx, r in enumerate(active, start=2):
        for col_idx, (name, _) in enumerate(COLUMNS, start=1):
            val = r.get(name, "-")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = tier_fills.get(r["Prospect Tier"], tier_fills["-"])
            if val in ("Not found", "-"):
                cell.font = Font(italic=True, color="9CA3AF")
    for i in range(2, len(active) + 2):
        ws.row_dimensions[i].height = 140

    ws2 = wb.create_sheet("Skipped")
    ws2.freeze_panes = "A2"
    for col_idx, (name, width) in enumerate(SKIP_COLUMNS, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[1].height = 28

    for row_idx, r in enumerate(skipped, start=2):
        values = [
            r.get("Domain", ""), r.get("Article URL", ""), r.get("Article Title", ""),
            extract_skip_reason(r), r.get("Source Engines", "-"), r.get("Why This Prospect", "-"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = skip_fill
    for i in range(2, len(skipped) + 2):
        ws2.row_dimensions[i].height = 60

    out_path = f"{base}.xlsx"
    wb.save(out_path)
    print(f"Wrote xlsx → {out_path}")
    print(f"  Outreach: {len(active)} rows × {len(COLUMNS)} cols")
    print(f"  Skipped:  {len(skipped)} rows × {len(SKIP_COLUMNS)} cols")

    # Metadata sidecar
    try:
        runner_meta = json.load(open(f"{base}_run_metadata.json"))
    except FileNotFoundError:
        runner_meta = {}

    tier_counts = Counter(r["Prospect Tier"] for r in rows if r["Outreach Status"] != "Skip")
    skip_counts = Counter()
    for r in rows:
        if r["Outreach Status"] == "Skip":
            m = re.search(r'SKIP[^:]*:\s*([^|]+)', r.get("Notes", "") or "")
            if m:
                skip_counts[m.group(1).strip()] += 1

    metadata = {
        "runId": runner_meta.get("runId", ""),
        "actorId": runner_meta.get("actorId", "apify/link-prospecting-tool"),
        "startedAt": runner_meta.get("startedAt", ""),
        "finishedAt": runner_meta.get("finishedAt", ""),
        "status": runner_meta.get("status", ""),
        "datasetIds": runner_meta.get("datasetIds", {}),
        "campaign": {
            "goal": cfg.get("goal"),
            "brand": cfg.get("brand", {}).get("name"),
            "userContentUrl": cfg.get("user_url"),
            "ownDomains": cfg.get("own_domains"),
            "competitorDomains": cfg.get("competitor_domains"),
            "alreadyPitchedDomains": cfg.get("already_pitched_domains", []),
            "partnershipType": cfg.get("partnership_type"),
            "brandVoice": cfg.get("brand_voice"),
        },
        "tierCounts": dict(tier_counts),
        "skipCounts": dict(skip_counts),
        "totalRows": len(rows),
        "activeRows": len(active),
        "skippedRows": len(skipped),
    }
    final_meta_path = f"{base}_run_metadata_final.json"
    with open(final_meta_path, "w") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    print(f"Wrote metadata sidecar → {final_meta_path}")
    print(f"\nFinal: total={metadata['totalRows']} active={metadata['activeRows']} skipped={metadata['skippedRows']}")
    print(f"Tier breakdown: {dict(tier_counts)}")


if __name__ == "__main__":
    main()
